from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import os
COEFS = {1: 0.3, 2: 0.4, 3: 0.5}
TITLES_STRENGTH = {1: "Strength_A", 2: "Strength_B", 3: "Strength_C"}

wb = Workbook()
ws = wb.active

class User:
    def __init__(self, name, weight):
        self.weight = weight
        self.level = 0
        self.titles_unlocked = []
        self.name = name
        if os.path.isfile(f'{name}.xlsx') == True:
            self.profile = load_workbook(f'{name}.xlsx')
        else:
            wb.title = name
            ws.append(['Credit card info','Amount', 'Donation', 'Total Contribuations'])
            wb.save(f'{name}.xlsx')
        class Strength:
            def __init__(self, user: User, weight_lifted):
                self.user = user
                self.weight_lifted = weight_lifted
            def levelStrength(self):
                goal_weight = self.user.weight * COEFS[self.user.level]
                ratio = self.weight_lifted / goal_weight
                if ratio >= 1:
                    self.user.level += 1
                    self.user.titles_unlocked.append(TITLES_STRENGTH[self.user.level])

