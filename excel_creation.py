from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

def creation(name):
    if os.path.isfile(f'Data/{name}.xlsx') == True:
        lw = load_workbook(f'Data/{name}.xlsx')
    else:
        wb.title = name
        ws.append([''])
        wb.save(f'{name}.xlsx')

def

