from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as char
import os

headings = ['Time', 'Lv', 'Strength', 'Endurance', 'Agility', 'Bench Press', 'Dead Lift', 'Squat']
directory = {'Time': [], 'Lv': [], 'Strength': [], 'Endurance': [], 'Agility': [], 'Bench Press': [], 'Dead Lift': [], 'Squat': []}
def creation(name):
    wb = Workbook()
    ws = wb.active
    if os.path.isfile(f'Data/{name}.xlsx') == True:
        wb = load_workbook(f'Data/{name}.xlsx')
        sh = wb['Sheet1']
        def check_cells():
            for section in range(1,8):
                for col in sh[char(section)]:
                    if col.value == headings[section-1]:
                        continue
                    else:
                        directory[headings[section-1]].append(col.value)
            print(directory)
        def data_entry(time, lv, stre, endr, agil, bp, dl, sq):
            check_cells()
            for col in range(1,8):
                for rows in range(len(directory[]))
    else:
        wb.title = name
        ws.append(headings)
        wb.save(f'{name}.xlsx')


# def insert(time, strength, endurance, lv, bp, dl, sq):
#     ws['A' + ]
creation('Andre')
#print(directory)