from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as char

from datetime import datetime
import os

headings = ['Time', 'Lv', 'Strength', 'Bench Press', 'Dead Lift', 'Squat']
wb = Workbook()
ws = wb.active
class fitQuest:
    def __init__(self, name, category):
        self.name = name
        self.cat = category
        self.stats = []
        if os.path.isfile(f'Data/{name}.xlsx') == False:
            wb.title = name
            ws.append(headings)
            wb.save(f'Data/{name}.xlsx')
        self.load = load_workbook(f'Data/{name}.xlsx')
        self.quest = self.load.active
        self.check_cells()
    def pos(self, rows):
        if self.cat == 'Time':
            return char(1) + str(rows+2)
        elif self.cat == 'Lv':
            return char(2) + str(rows+2)
        elif self.cat == 'Strength':
            return char(3) + str(rows+2)
        elif self.cat == 'Bench Press':
            return char(4) + str(rows+2)
        elif self.cat == 'Dead Lift':
            return char(5) + str(rows+2)
        else:
            return char(6) + str(rows+2)
    def check_cells(self):
        for row in range(1000):
            if self.quest[self.pos(row)].value == None:
                break
            else:
                self.stats.append(self.quest[self.pos(row)].value)

    def add_entry(self, num):
        self.stats.append(num)
        for row in range(len(self.stats)):
            self.quest[self.pos(row)] = self.stats[row]
        self.load.save(f'Data/{self.name}.xlsx')





