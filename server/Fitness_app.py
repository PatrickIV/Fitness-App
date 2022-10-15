from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

def creation():

Filepath = "Data/Filename.xlsx"
wb.title = 'Filename'
wb.save(Filepath)

